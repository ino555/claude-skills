"""
save_to_excel.py — streaming-film-finder skill helper

Usage:
    python save_to_excel.py '<json_array>' '<output_path>'

JSON array format:
[
  {
    "rank": 1,
    "title": "Film Title",
    "title_tr": "Turkce Baslik (optional)",
    "genres": "Action, Thriller",
    "platforms": "Netflix, Prime Video",
    "added_date": "2026-04-15",
    "imdb_score": 7.4,
    "rt_score": 82,
    "review_summary": "Short review summary.",
    "imdb_url": "https://www.imdb.com/title/tt1234567/",
    "watch_url": "https://www.netflix.com/title/12345"
  },
  ...
]
"""

import sys
import json
from datetime import date

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not found. Installing...", file=sys.stderr)
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


def create_excel(films: list, output_path: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Top 5 Streaming Films"

    today = date.today().strftime("%d %B %Y")

    HEADER_FILL = PatternFill("solid", fgColor="1E4D8C")
    GOLD_FILL   = PatternFill("solid", fgColor="FFD700")
    ALT_FILL    = PatternFill("solid", fgColor="EBF2FF")
    WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")

    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    TITLE_FONT  = Font(bold=True, size=14, color="1E4D8C")
    LINK_FONT   = Font(color="0563C1", underline="single", size=10)
    BODY_FONT   = Font(size=10)

    thin = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    ws.merge_cells("A1:J1")
    ws["A1"] = f"En Iyi 5 Streaming Filmi  --  {today}"
    ws["A1"].font      = TITLE_FONT
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    headers = [
        "#", "Film Adi", "Tur", "Platform(lar)",
        "Eklenme Tarihi", "IMDB", "RT%",
        "Yorum Ozeti", "IMDB Linki", "Izle"
    ]
    col_widths = [4, 28, 18, 20, 16, 8, 6, 40, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[2].height = 22

    for i, film in enumerate(films):
        row = i + 3
        imdb = float(film.get("imdb_score") or 0)
        fill = GOLD_FILL if imdb >= 8.0 else (ALT_FILL if i % 2 == 0 else WHITE_FILL)

        title_display = film.get("title", "")
        if film.get("title_tr"):
            title_display += f"\n({film['title_tr']})"

        rt_val = film.get("rt_score")
        rt_display = f"{rt_val}%" if rt_val else "N/A"
        imdb_display = f"{imdb:.1f}" if imdb else "N/A"

        values = [
            film.get("rank", i + 1),
            title_display,
            film.get("genres", ""),
            film.get("platforms", ""),
            film.get("added_date", ""),
            imdb_display,
            rt_display,
            film.get("review_summary", ""),
            "",
            "",
        ]

        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.fill      = fill
            cell.border    = BORDER
            cell.font      = BODY_FONT
            cell.alignment = CENTER if col in (1, 5, 6, 7) else LEFT

        imdb_url  = film.get("imdb_url", "")
        watch_url = film.get("watch_url", "")

        if imdb_url:
            link_cell = ws.cell(row=row, column=9, value="IMDB ->")
            link_cell.hyperlink  = imdb_url
            link_cell.font       = LINK_FONT
            link_cell.fill       = fill
            link_cell.border     = BORDER
            link_cell.alignment  = CENTER

        if watch_url:
            watch_cell = ws.cell(row=row, column=10, value="Izle ->")
            watch_cell.hyperlink = watch_url
            watch_cell.font      = LINK_FONT
            watch_cell.fill      = fill
            watch_cell.border    = BORDER
            watch_cell.alignment = CENTER

        ws.row_dimensions[row].height = 42

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:J{2 + len(films)}"

    legend_row = 3 + len(films) + 1
    ws.merge_cells(f"A{legend_row}:J{legend_row}")
    legend_cell = ws.cell(
        row=legend_row,
        column=1,
        value="Altin arka plan = IMDB 8.0+   |   Tum veriler arastirma tarihinde gecerlidir."
    )
    legend_cell.font      = Font(italic=True, size=9, color="666666")
    legend_cell.alignment = LEFT

    wb.save(output_path)
    print(f"Excel saved: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python save_to_excel.py '<json>' '<output.xlsx>'")
        sys.exit(1)

    raw_json    = sys.argv[1]
    output_path = sys.argv[2]

    try:
        films = json.loads(raw_json)
    except json.JSONDecodeError as e:
        print(f"Invalid JSON: {e}", file=sys.stderr)
        sys.exit(1)

    create_excel(films, output_path)
